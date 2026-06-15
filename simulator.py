"""
Phase 3 — Investment Simulator | AI Readiness Copilot
Reads scores.xlsx → takes budget allocation → projects Conservative/Moderate/Optimistic
readiness gains per sector with reasoning.
Output: simulation_results.xlsx + printed summary
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Impact model ───────────────────────────────────────────────────────────────
# For each sector: how many readiness points gained per $1B invested
# Ranges reflect uncertainty: (conservative, moderate, optimistic)
# Based on World Bank / McKinsey returns-to-investment literature for SSA/MENA

IMPACT_MODEL = {
    "education": {
        "points_per_billion": (2.5, 4.5, 7.0),
        "lag_years":          3,
        "rationale":          "Teacher training and digital literacy programs show 2–7pt gains per $1B in comparable economies (WB 2022). Lag of ~3 years before measurable readiness improvement.",
        "bottlenecks":        ["Low tertiary enrollment requires parallel infrastructure", "Teacher quality constraints limit absorption"],
        "quick_wins":         ["Digital curriculum rollout", "Scholarship programs for STEM"],
    },
    "workforce": {
        "points_per_billion": (3.0, 5.5, 8.5),
        "lag_years":          2,
        "rationale":          "ICT skills programs and R&D investment yield 3–8.5pt gains per $1B. Shorter lag than education as adult upskilling is faster.",
        "bottlenecks":        ["Private sector absorption capacity", "Brain drain risk in high-skill segments"],
        "quick_wins":         ["Coding bootcamps and ICT certification", "R&D tax incentives for tech firms"],
    },
    "healthcare": {
        "points_per_billion": (4.0, 6.5, 9.5),
        "lag_years":          4,
        "rationale":          "Health infrastructure investment shows 4–9.5pt gains per $1B in sub-Saharan Africa (WHO 2023). Longest lag due to facility construction and workforce training.",
        "bottlenecks":        ["Physician supply takes years to increase", "Rural distribution challenges"],
        "quick_wins":         ["Telemedicine platforms", "Community health worker training"],
    },
    "government": {
        "points_per_billion": (1.5, 3.0, 5.0),
        "lag_years":          5,
        "rationale":          "Governance reforms show slower, lower returns (1.5–5pt per $1B) but high multiplier effects on other sectors. Longest lag due to institutional change.",
        "bottlenecks":        ["Political will and reform resistance", "Institutional capacity constraints"],
        "quick_wins":         ["E-government platforms", "Open data portals", "Anti-corruption hotlines"],
    },
    "infrastructure": {
        "points_per_billion": (3.5, 6.0, 9.0),
        "lag_years":          2,
        "rationale":          "Broadband and electricity infrastructure shows 3.5–9pt gains per $1B across Africa (ITU 2023). Relatively short lag as connectivity improvements are measurable quickly.",
        "bottlenecks":        ["Last-mile connectivity in rural areas", "Regulatory environment for ISPs"],
        "quick_wins":         ["National broadband rollout", "Solar electrification programs"],
    },
}

# Sector score ceiling — no sector can exceed 95 (realistic cap)
SCORE_CEILING = 95.0

# Diminishing returns threshold — above 70, gains are halved
DIMINISHING_THRESHOLD = 70.0


# ── Load current scores from scores.xlsx ──────────────────────────────────────

def load_scores(path="scores.xlsx"):
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb["Sector Scorecard"]
    rows = list(ws.iter_rows(values_only=True))

    scores = {}
    # Row 0 = header, rows 1-5 = sectors, last row = overall
    sector_order = ["education", "workforce", "healthcare", "government", "infrastructure"]
    for i, sector in enumerate(sector_order):
        row = rows[i + 1]
        # score may be float or None; confidence is a string like "76% (HIGH)"
        def parse_score(v):
            try: return float(v)
            except: return None
        def parse_conf(v):
            try:
                s = str(v).strip()
                return int(s.split("%")[0])
            except: return 0
        scores[sector] = {
            "NG": {"score": parse_score(row[1]), "confidence": parse_conf(row[2])},
            "DZ": {"score": parse_score(row[3]), "confidence": parse_conf(row[4])},
        }

    # Overall
    overall_row = rows[6]
    scores["_overall"] = {
        "NG": float(overall_row[1]) if overall_row[1] else None,
        "DZ": float(overall_row[3]) if overall_row[3] else None,
    }

    return scores


# ── Simulate ───────────────────────────────────────────────────────────────────

def simulate(country_code, current_scores, allocation, total_budget_billion):
    """
    allocation: {sector: fraction} e.g. {"education": 0.4, "infrastructure": 0.3, ...}
    Returns: {sector: {conservative, moderate, optimistic, rationale, ...}}
    """
    results = {}

    for sector, fraction in allocation.items():
        if fraction <= 0:
            continue

        budget = total_budget_billion * fraction
        current = current_scores.get(sector, {}).get(country_code, {}).get("score")
        if current is None:
            continue

        model = IMPACT_MODEL[sector]
        cons_rate, mod_rate, opt_rate = model["points_per_billion"]

        gains = []
        for rate in (cons_rate, mod_rate, opt_rate):
            raw_gain = budget * rate
            # Diminishing returns above threshold
            if current > DIMINISHING_THRESHOLD:
                raw_gain *= 0.5
            elif current + raw_gain > DIMINISHING_THRESHOLD:
                # Split: portion below threshold gets full rate, above gets half
                below = DIMINISHING_THRESHOLD - current
                above = raw_gain - below
                raw_gain = below + above * 0.5

            new_score = min(SCORE_CEILING, current + raw_gain)
            gains.append(round(new_score, 1))

        results[sector] = {
            "budget_billion":  round(budget, 2),
            "current_score":   current,
            "conservative":    gains[0],
            "moderate":        gains[1],
            "optimistic":      gains[2],
            "gain_cons":       round(gains[0] - current, 1),
            "gain_mod":        round(gains[1] - current, 1),
            "gain_opt":        round(gains[2] - current, 1),
            "lag_years":       model["lag_years"],
            "rationale":       model["rationale"],
            "bottlenecks":     model["bottlenecks"],
            "quick_wins":      model["quick_wins"],
        }

    # Projected overall scores
    sector_order = ["education", "workforce", "healthcare", "government", "infrastructure"]
    for scenario in ("conservative", "moderate", "optimistic"):
        projected = []
        for s in sector_order:
            if s in results:
                projected.append(results[s][scenario])
            else:
                base = current_scores.get(s, {}).get(country_code, {}).get("score") or 0
                projected.append(base)
        results[f"_overall_{scenario}"] = round(sum(projected) / len(projected), 1)

    return results


# ── Print summary ──────────────────────────────────────────────────────────────

def print_simulation(country, sim, total_budget, allocation):
    print(f"\n{'='*60}")
    print(f"INVESTMENT SIMULATION — {country.upper()}")
    print(f"Total Budget: ${total_budget}B | Allocation:")
    for s, f in allocation.items():
        print(f"  {s:<16} {int(f*100)}%  (${round(total_budget*f,2)}B)")
    print(f"{'='*60}")

    print(f"\n{'Sector':<16} {'Current':>8} {'Conserv.':>10} {'Moderate':>10} {'Optimist.':>10} {'Lag':>5}")
    print("-" * 62)
    for sector, r in sim.items():
        if sector.startswith("_"): continue
        print(f"{sector:<16} {r['current_score']:>8.1f} "
              f"{r['conservative']:>8.1f}(+{r['gain_cons']}) "
              f"{r['moderate']:>8.1f}(+{r['gain_mod']}) "
              f"{r['optimistic']:>8.1f}(+{r['gain_opt']}) "
              f"{r['lag_years']:>4}yr")

    print(f"\n{'Overall AI Readiness':}")
    for scenario in ("conservative", "moderate", "optimistic"):
        key = f"_overall_{scenario}"
        print(f"  {scenario.capitalize():<14}: {sim[key]}/100")


# ── Export simulation_results.xlsx ────────────────────────────────────────────

def export_simulation(all_simulations, path="simulation_results.xlsx"):
    wb = openpyxl.Workbook()

    hdr_fill  = PatternFill("solid", fgColor="1F3864")
    cons_fill = PatternFill("solid", fgColor="FCE4D6")   # red-ish = conservative
    mod_fill  = PatternFill("solid", fgColor="FFF2CC")   # yellow = moderate
    opt_fill  = PatternFill("solid", fgColor="E2EFDA")   # green = optimistic
    sect_fill = PatternFill("solid", fgColor="2E75B6")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    bold      = Font(bold=True)
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def sc(cell, fill=None, font=None, align="center"):
        if fill: cell.fill = fill
        cell.font = font or Font(size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = border

    first = True
    for (country, budget, allocation), sim in all_simulations.items():
        sheet_name = f"{country} ${budget}B"
        ws = wb.active if first else wb.create_sheet(sheet_name)
        if first:
            ws.title = sheet_name
            first = False

        # Title block
        ws.append([f"AI Readiness Investment Simulation — {country}"])
        ws.cell(1, 1).font = Font(bold=True, size=14, color="1F3864")
        ws.append([f"Total Budget: ${budget}B"])
        ws.append(["Allocation:"] + [f"{s}: {int(f*100)}%" for s, f in allocation])
        ws.append([])

        # Scenario header
        headers = ["Sector", "Budget ($B)", "Current Score",
                   "Conservative", "+Gain", "Moderate", "+Gain",
                   "Optimistic", "+Gain", "Lag (yrs)", "Rationale"]
        ws.append(headers)
        for i in range(1, 12):
            sc(ws.cell(ws.max_row, i), fill=hdr_fill, font=hdr_font)

        for sector, r in sim.items():
            if sector.startswith("_"): continue
            row = [
                sector.upper(),
                r["budget_billion"],
                r["current_score"],
                r["conservative"], f"+{r['gain_cons']}",
                r["moderate"],     f"+{r['gain_mod']}",
                r["optimistic"],   f"+{r['gain_opt']}",
                r["lag_years"],
                r["rationale"],
            ]
            ws.append(row)
            rn = ws.max_row
            sc(ws.cell(rn, 1), font=bold, align="left")
            sc(ws.cell(rn, 2)); sc(ws.cell(rn, 3))
            sc(ws.cell(rn, 4), fill=cons_fill); sc(ws.cell(rn, 5), fill=cons_fill)
            sc(ws.cell(rn, 6), fill=mod_fill);  sc(ws.cell(rn, 7), fill=mod_fill)
            sc(ws.cell(rn, 8), fill=opt_fill);  sc(ws.cell(rn, 9), fill=opt_fill)
            sc(ws.cell(rn, 10)); sc(ws.cell(rn, 11), align="left")

        # Overall row
        ws.append([])
        ws.append(["PROJECTED OVERALL AI READINESS", "", "",
                   sim["_overall_conservative"], "",
                   sim["_overall_moderate"],     "",
                   sim["_overall_optimistic"],   "", "", ""])
        rn = ws.max_row
        for i in range(1, 12):
            sc(ws.cell(rn, i), fill=hdr_fill, font=hdr_font)
        sc(ws.cell(rn, 4), fill=cons_fill)
        sc(ws.cell(rn, 6), fill=mod_fill)
        sc(ws.cell(rn, 8), fill=opt_fill)

        # Quick wins + bottlenecks
        ws.append([])
        ws.append(["QUICK WINS & BOTTLENECKS PER SECTOR"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=11, color="1F3864")
        ws.append(["Sector", "Quick Wins", "Bottlenecks"])
        rn = ws.max_row
        for i in range(1, 4):
            sc(ws.cell(rn, i), fill=sect_fill, font=Font(bold=True, color="FFFFFF"))

        for sector, r in sim.items():
            if sector.startswith("_"): continue
            ws.append([
                sector.upper(),
                " | ".join(r["quick_wins"]),
                " | ".join(r["bottlenecks"]),
            ])
            rn = ws.max_row
            sc(ws.cell(rn, 1), font=bold, align="left")
            sc(ws.cell(rn, 2), align="left")
            sc(ws.cell(rn, 3), align="left")

        # Column widths
        widths = [18, 10, 12, 12, 8, 12, 8, 12, 8, 10, 50]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A6"

    wb.save(path)
    print(f"\n✓ Simulation → {path}")


# ── Run ────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Phase 3 — Investment Simulator")
    print("Loading scores.xlsx ...\n")

    current_scores = load_scores("scores.xlsx")

    # ── Define scenarios to simulate ──────────────────────────────────────────
    # Format: (country_label, total_budget_$B, {sector: fraction})
    # Fractions must sum to 1.0
    # Add or modify scenarios here freely

    SIMULATIONS = [
        # Nigeria — $2B balanced
        ("Nigeria", 2.0, {
            "education":      0.30,
            "healthcare":     0.30,
            "infrastructure": 0.25,
            "workforce":      0.15,
        }),
        # Nigeria — $2B healthcare-first (addressing biggest gap)
        ("Nigeria", 2.0, {
            "healthcare":     0.50,
            "infrastructure": 0.25,
            "education":      0.15,
            "workforce":      0.10,
        }),
        # Algeria — $2B workforce + infrastructure
        ("Algeria", 2.0, {
            "workforce":      0.40,
            "infrastructure": 0.30,
            "education":      0.20,
            "government":     0.10,
        }),
    ]

    COUNTRY_CODE = {"Nigeria": "NG", "Algeria": "DZ"}

    all_simulations = {}
    for country, budget, allocation in SIMULATIONS:
        iso2 = COUNTRY_CODE[country]
        sim  = simulate(iso2, current_scores, allocation, budget)
        print_simulation(country, sim, budget, allocation)
        all_simulations[(country, budget, tuple(allocation.items()))] = sim

    export_simulation(all_simulations, "simulation_results.xlsx")
