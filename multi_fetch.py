"""
Multi-Source Data Fetcher — AI Readiness Copilot
Sources: World Bank → ITU → UNESCO UIS → WHO GHO (in priority order)
Countries: Nigeria (NG / NGA), Algeria (DZ / DZA)

Strategy: fetch World Bank first, then fill MISSING slots from secondary sources.
Every indicator record carries its source so the Responsible AI Center can show provenance.
"""

import requests
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ─────────────────────────────────────────────────────────────────────

COUNTRIES = {
    "NG": {"name": "Nigeria",  "iso3": "NGA", "who": "NGA", "itu": "NGA"},
    "DZ": {"name": "Algeria",  "iso3": "DZA", "who": "DZA", "itu": "DZA"},
}

STALE_YEARS = 5
NOW_YEAR    = datetime.now().year

# ── Indicator definitions ──────────────────────────────────────────────────────
# Each indicator lists sources in priority order.
# "wb"  → World Bank REST API
# "itu" → ITU datahub API
# "unesco" → UNESCO UIS API
# "who" → WHO GHO API

INDICATORS = {

    # ── EDUCATION ──────────────────────────────────────────────────────────────
    "education": [
        {
            "id": "adult_literacy",
            "label": "Adult literacy rate (%)",
            "sources": [
                {"type": "wb",      "code": "SE.ADT.LITR.ZS"},
                {"type": "unesco",  "code": "LR.AG15T99"},
            ],
        },
        {
            "id": "edu_spend_gdp",
            "label": "Gov education spend (% GDP)",
            "sources": [
                {"type": "wb",      "code": "SE.XPD.TOTL.GD.ZS"},
                {"type": "unesco",  "code": "XGDP.FSGOV"},
            ],
        },
        {
            "id": "tertiary_enrollment",
            "label": "Tertiary enrollment rate (%)",
            "sources": [
                {"type": "wb",      "code": "SE.TER.ENRR"},
                {"type": "unesco",  "code": "ROFST.H.3"},
            ],
        },
        {
            "id": "primary_completion",
            "label": "Primary completion rate (%)",
            "sources": [
                {"type": "wb",      "code": "SE.PRM.CMPT.ZS"},
                {"type": "unesco",  "code": "CR.1"},
            ],
        },
        {
            "id": "trained_teachers",
            "label": "Trained teachers, primary (%)",
            "sources": [
                {"type": "unesco",  "code": "TRTP.1"},
                {"type": "wb",      "code": "SE.PRM.TCAQ.ZS"},
            ],
        },
    ],

    # ── WORKFORCE ──────────────────────────────────────────────────────────────
    "workforce": [
        {
            "id": "labor_force",
            "label": "Total labor force",
            "sources": [{"type": "wb", "code": "SL.TLF.TOTL.IN"}],
        },
        {
            "id": "unemployment",
            "label": "Unemployment rate (%)",
            "sources": [{"type": "wb", "code": "SL.UEM.TOTL.ZS"}],
        },
        {
            "id": "rd_spend",
            "label": "R&D expenditure (% GDP)",
            "sources": [
                {"type": "wb",      "code": "GB.XPD.RSDV.GD.ZS"},
                {"type": "unesco",  "code": "XGDP.RNDP"},
            ],
        },
        {
            "id": "ict_skills",
            "label": "Population with ICT skills (%)",
            "sources": [
                {"type": "itu",     "code": "i99H"},   # ITU ICT skills indicator
            ],
        },
    ],

    # ── HEALTHCARE ─────────────────────────────────────────────────────────────
    "healthcare": [
        {
            "id": "health_spend_gdp",
            "label": "Health expenditure (% GDP)",
            "sources": [
                {"type": "wb",  "code": "SH.XPD.CHEX.GD.ZS"},
                {"type": "who", "code": "GHED_CHE_pc_PPP_SHA2011"},
            ],
        },
        {
            "id": "hospital_beds",
            "label": "Hospital beds per 1,000",
            "sources": [
                {"type": "wb",  "code": "SH.MED.BEDS.ZS"},
                {"type": "who", "code": "HWF_0001"},
            ],
        },
        {
            "id": "under5_mortality",
            "label": "Under-5 mortality rate",
            "sources": [
                {"type": "wb",  "code": "SH.DYN.MORT"},
                {"type": "who", "code": "MDG_0000000007"},
            ],
        },
        {
            "id": "physicians",
            "label": "Physicians per 1,000",
            "sources": [
                {"type": "wb",  "code": "SH.MED.PHYS.ZS"},
                {"type": "who", "code": "HWF_0001"},
            ],
        },
        {
            "id": "uhc_index",
            "label": "UHC service coverage index",
            "sources": [
                {"type": "wb",  "code": "SH.UHC.SRVS.CV.XD"},
                {"type": "who", "code": "UHC_INDEX_REPORTED"},
            ],
        },
    ],

    # ── GOVERNMENT ─────────────────────────────────────────────────────────────
    # WGI (GE.EST etc.) requires separate API key — filled by IIAG merge instead
    "government": [],

    # ── INFRASTRUCTURE ─────────────────────────────────────────────────────────
    "infrastructure": [
        {
            "id": "internet_users",
            "label": "Internet users (% population)",
            "sources": [
                {"type": "wb",  "code": "IT.NET.USER.ZS"},
                {"type": "itu", "code": "i99B"},
            ],
        },
        {
            "id": "mobile_subs",
            "label": "Mobile subscriptions per 100",
            "sources": [
                {"type": "wb",  "code": "IT.CEL.SETS.P2"},
                {"type": "itu", "code": "i271"},
            ],
        },
        {
            "id": "electricity_access",
            "label": "Access to electricity (%)",
            "sources": [{"type": "wb", "code": "EG.ELC.ACCS.ZS"}],
        },
        {
            "id": "broadband_subs",
            "label": "Fixed broadband subscriptions per 100",
            "sources": [
                {"type": "wb",  "code": "IT.NET.BBND.P2"},
                {"type": "itu", "code": "i4213"},
            ],
        },
        {
            "id": "mobile_broadband",
            "label": "Mobile broadband subscriptions per 100",
            "sources": [
                {"type": "itu", "code": "i4214"},
                {"type": "wb",  "code": "IT.MOB.BNDW.P2"},
            ],
        },
    ],
}


# ── Source fetchers ────────────────────────────────────────────────────────────

def _flag(year_str):
    try:
        age = NOW_YEAR - int(str(year_str)[:4])
        return "stale" if age >= STALE_YEARS else "ok"
    except:
        return "ok"


def fetch_wb(code, country_iso2, source_param=None):
    """World Bank REST API."""
    sp = f"&source={source_param}" if source_param else ""
    url = (
        f"https://api.worldbank.org/v2/country/{country_iso2}/indicator/{code}"
        f"?format=json&date=2010:2025&per_page=15&mrv=1{sp}"
    )
    try:
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        data = r.json()
        if len(data) < 2 or not data[1]:
            return None, None
        for entry in data[1]:
            if entry.get("value") is not None:
                return entry["value"], str(entry["date"])
    except Exception as e:
        print(f"    [WB error] {code}/{country_iso2}: {e}")
    return None, None


def fetch_itu(code, country_iso3):
    """ITU Datahub API."""
    url = (
        f"https://datahub.itu.int/api/service/indicator/data"
        f"?indicator={code}&entity={country_iso3}&format=json"
    )
    try:
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        data = r.json()
        # ITU returns list of {year, value} — pick latest non-null
        entries = data.get("data", []) or data if isinstance(data, list) else []
        entries_sorted = sorted(entries, key=lambda x: x.get("year", 0), reverse=True)
        for e in entries_sorted:
            v = e.get("value") or e.get("Value")
            y = e.get("year") or e.get("Year")
            if v is not None:
                return float(v), str(y)
    except Exception as e:
        print(f"    [ITU error] {code}/{country_iso3}: {e}")
    return None, None


def fetch_unesco(code, country_iso3):
    """UNESCO UIS API (no key required for most indicators)."""
    url = (
        f"https://api.uis.unesco.org/api/public/data/indicators"
        f"?indicator={code}&geoUnit={country_iso3}&version=20231&format=json"
    )
    try:
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        data = r.json()
        records = data.get("data", [])
        records_sorted = sorted(records, key=lambda x: x.get("year", 0), reverse=True)
        for rec in records_sorted:
            v = rec.get("value")
            y = rec.get("year")
            if v is not None:
                return float(v), str(y)
    except Exception as e:
        print(f"    [UNESCO error] {code}/{country_iso3}: {e}")
    return None, None


def fetch_who(code, country_iso3):
    """WHO Global Health Observatory API."""
    url = (
        f"https://ghoapi.azureedge.net/api/{code}"
        f"?$filter=SpatialDim eq '{country_iso3}'"
        f"&$orderby=TimeDim desc&$top=5"
    )
    try:
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        data = r.json()
        for entry in data.get("value", []):
            v = entry.get("NumericValue")
            y = entry.get("TimeDim")
            if v is not None:
                return float(v), str(y)
    except Exception as e:
        print(f"    [WHO error] {code}/{country_iso3}: {e}")
    return None, None


# ── Dispatch ───────────────────────────────────────────────────────────────────

def fetch_from_source(src, country_iso2, country_iso3):
    """Try one source definition. Returns (value, year, source_type) or (None, None, None)."""
    t = src["type"]
    code = src["code"]

    if t == "wb":
        v, y = fetch_wb(code, country_iso2, src.get("source_param"))
    elif t == "itu":
        v, y = fetch_itu(code, country_iso3)
    elif t == "unesco":
        v, y = fetch_unesco(code, country_iso3)
    elif t == "who":
        v, y = fetch_who(code, country_iso3)
    else:
        return None, None, None

    if v is not None:
        return v, y, t
    return None, None, None


def fetch_indicator(indicator, country_iso2, country_iso3):
    """Try each source in order, return first hit."""
    for src in indicator["sources"]:
        v, y, source_type = fetch_from_source(src, country_iso2, country_iso3)
        if v is not None:
            return {
                "id":       indicator["id"],
                "label":    indicator["label"],
                "value":    v,
                "year":     y,
                "source":   source_type,
                "source_code": src["code"],
                "missing":  False,
                "flag":     _flag(y),
            }

    # All sources exhausted
    return {
        "id":          indicator["id"],
        "label":       indicator["label"],
        "value":       None,
        "year":        None,
        "source":      None,
        "source_code": None,
        "missing":     True,
        "flag":        "no_data",
    }


# ── Main fetch ─────────────────────────────────────────────────────────────────

def fetch_all():
    results = {}

    for iso2, meta in COUNTRIES.items():
        name    = meta["name"]
        iso3    = meta["iso3"]
        results[iso2] = {"name": name, "sectors": {}}

        for sector, indicators in INDICATORS.items():
            results[iso2]["sectors"][sector] = {}
            print(f"\n[{name}] {sector.upper()}")

            for ind in indicators:
                rec = fetch_indicator(ind, iso2, iso3)

                if not rec["missing"]:
                    stale = " ⚠ STALE" if rec["flag"] == "stale" else ""
                    print(f"  ✓ [{rec['source'].upper()}] {rec['label']}: {rec['value']} ({rec['year']}){stale}")
                else:
                    print(f"  ✗ MISSING  {rec['label']}")

                results[iso2]["sectors"][sector][rec["id"]] = rec

    return results


# ── Export ─────────────────────────────────────────────────────────────────────

def export_json(results, path="all_data.json"):
    with open(path, "w") as f:
        json.dump(results, f, indent=2)
    print(f"\n✓ JSON → {path}")


def export_excel(results, path="all_data.xlsx"):
    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_fill   = PatternFill("solid", fgColor="1F3864")   # dark blue
    sect_fill  = PatternFill("solid", fgColor="2E75B6")   # mid blue
    stale_fill = PatternFill("solid", fgColor="FFF2CC")   # yellow
    miss_fill  = PatternFill("solid", fgColor="FCE4D6")   # red-orange
    ok_fill    = PatternFill("solid", fgColor="E2EFDA")   # green

    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    sect_font  = Font(bold=True, color="FFFFFF", size=10)
    bold       = Font(bold=True)
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_cell(cell, fill=None, font=None, align="left", bold_=False):
        if fill:   cell.fill = fill
        if font:   cell.font = font
        elif bold_: cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = border

    # ── Sheet 1: All Data (flat table) ────────────────────────────────────────
    ws = wb.active
    ws.title = "All Data"

    headers = ["Country", "Country Code", "Sector", "Indicator ID",
               "Indicator Label", "Value", "Year", "Source",
               "Source Code", "Missing", "Flag"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        style_cell(cell, fill=hdr_fill, font=hdr_font, align="center")

    row_num = 2
    for iso2, country_data in results.items():
        for sector, indicators in country_data["sectors"].items():
            for ind_id, rec in indicators.items():
                ws.append([
                    country_data["name"], iso2, sector, ind_id,
                    rec["label"], rec["value"], rec["year"],
                    rec["source"], rec["source_code"],
                    rec["missing"], rec["flag"],
                ])
                # color-code by flag
                row_fill = miss_fill if rec["missing"] else (stale_fill if rec["flag"] == "stale" else ok_fill)
                for col_idx in range(1, len(headers)+1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    style_cell(cell, fill=row_fill)
                row_num += 1

    # Column widths
    widths = [12, 12, 14, 22, 40, 12, 8, 10, 18, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Per-country sector summary ───────────────────────────────────
    ws2 = wb.create_sheet("By Sector")
    ws2.append(["Sector", "Indicator", "Nigeria Value", "Nigeria Year",
                "Nigeria Flag", "Algeria Value", "Algeria Year", "Algeria Flag"])
    for col_idx in range(1, 9):
        style_cell(ws2.cell(row=1, column=col_idx), fill=hdr_fill, font=hdr_font, align="center")

    row_num2 = 2
    # Collect all unique sectors+indicators
    seen = {}
    for iso2, country_data in results.items():
        for sector, indicators in country_data["sectors"].items():
            for ind_id, rec in indicators.items():
                key = (sector, ind_id, rec["label"])
                if key not in seen:
                    seen[key] = {}
                seen[key][iso2] = rec

    current_sector = None
    for (sector, ind_id, label), countries in seen.items():
        # Sector header row
        if sector != current_sector:
            current_sector = sector
            ws2.append([sector.upper(), "", "", "", "", "", "", ""])
            for col_idx in range(1, 9):
                style_cell(ws2.cell(row=row_num2, column=col_idx), fill=sect_fill, font=sect_font)
            row_num2 += 1

        ng = countries.get("NG", {})
        dz = countries.get("DZ", {})
        ws2.append([
            "", label,
            ng.get("value"), ng.get("year"), ng.get("flag"),
            dz.get("value"), dz.get("year"), dz.get("flag"),
        ])
        for col_idx in range(1, 9):
            cell = ws2.cell(row=row_num2, column=col_idx)
            ng_flag = ng.get("flag", "")
            row_fill = miss_fill if ng.get("missing") else (stale_fill if ng_flag == "stale" else None)
            style_cell(cell, fill=row_fill)
        row_num2 += 1

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 40
    for col in ["C","D","E","F","G","H"]:
        ws2.column_dimensions[col].width = 16
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Coverage summary ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Coverage Summary")
    ws3.append(["Country", "Sector", "Total", "Fetched", "Missing", "Stale", "Coverage %"])
    for col_idx in range(1, 8):
        style_cell(ws3.cell(row=1, column=col_idx), fill=hdr_fill, font=hdr_font, align="center")

    row_num3 = 2
    for iso2, country_data in results.items():
        for sector, indicators in country_data["sectors"].items():
            total = len(indicators)
            missing = sum(1 for r in indicators.values() if r["missing"])
            stale   = sum(1 for r in indicators.values() if r.get("flag") == "stale")
            fetched = total - missing
            pct     = round(fetched / total * 100) if total else 0
            ws3.append([country_data["name"], sector, total, fetched, missing, stale, pct])
            pct_cell = ws3.cell(row=row_num3, column=7)
            pct_fill = ok_fill if pct >= 80 else (stale_fill if pct >= 60 else miss_fill)
            for col_idx in range(1, 8):
                style_cell(ws3.cell(row=row_num3, column=col_idx), fill=pct_fill if col_idx==7 else None)
            row_num3 += 1

    for col, w in zip(["A","B","C","D","E","F","G"], [14,16,8,10,10,8,12]):
        ws3.column_dimensions[col].width = w

    wb.save(path)
    print(f"✓ Excel → {path}  (sheets: All Data | By Sector | Coverage Summary)")


def print_summary(results):
    print("\n" + "="*55)
    print("FETCH SUMMARY")
    print("="*55)
    for iso2, country_data in results.items():
        total = missing = stale = 0
        by_source = {}
        for sector, indicators in country_data["sectors"].items():
            for rec in indicators.values():
                total += 1
                if rec["missing"]:
                    missing += 1
                else:
                    if rec["flag"] == "stale":
                        stale += 1
                    src = rec["source"]
                    by_source[src] = by_source.get(src, 0) + 1

        coverage = round((total - missing) / total * 100)
        src_breakdown = " | ".join(f"{k.upper()}:{v}" for k, v in sorted(by_source.items()))
        print(f"\n{country_data['name']}")
        print(f"  Coverage : {total-missing}/{total} ({coverage}%)")
        print(f"  Stale    : {stale} indicators (>{STALE_YEARS}yr old)")
        print(f"  Sources  : {src_breakdown}")
        if missing > 0:
            print(f"  Missing  : {missing} → add manually from NBS/ONS PDFs or mark synthetic")


# ── Run ────────────────────────────────────────────────────────────────────────

# ── Mo Ibrahim IIAG loader ─────────────────────────────────────────────────────
# IIAG has no public API — download manually from:
# https://iiag.online/downloads.html → "2024 IIAG: Data"
# Save as iiag_2024.xlsx in same folder as this script.

IIAG_INDICATORS = {
    "government": [
        {"id": "iiag_overall",        "label": "IIAG Overall Governance score",            "col": "GOVERNANCE"},
        {"id": "iiag_srol",           "label": "IIAG Security & Rule of Law",              "col": "SROL"},
        {"id": "iiag_rule_of_law",    "label": "IIAG Rule of Law & Justice",               "col": "ROLJUS"},
        {"id": "iiag_accountability", "label": "IIAG Accountability & Transparency",       "col": "ACCTRANS"},
    ],
    "infrastructure": [
        {"id": "iiag_infra",          "label": "IIAG Infrastructure score",               "col": "INFR"},
    ],
    "education": [
        {"id": "iiag_human_dev",      "label": "IIAG Human Development score",            "col": "HD"},
        {"id": "iiag_education",      "label": "IIAG Education score",                    "col": "EDUC"},
    ],
    "healthcare": [
        {"id": "iiag_health",         "label": "IIAG Health score",                       "col": "HEALTH"},
    ],
    "workforce": [
        {"id": "iiag_econ_opp",       "label": "IIAG Foundations for Economic Opportunity", "col": "FEO"},
        {"id": "iiag_business_env",   "label": "IIAG Business Environment score",         "col": "BUSENV"},
    ],
}

IIAG_ISO3 = {"NG": "NGA", "DZ": "DZA"}


def load_iiag(path=None):
    import os
    if path is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        home = os.path.expanduser("~")
        candidates = [
            os.path.join(script_dir, "2024-IIAG-scores.xlsx"),
            os.path.join(script_dir, "iiag_2024.xlsx"),
            os.path.join(os.getcwd(), "2024-IIAG-scores.xlsx"),
            os.path.join(os.getcwd(), "iiag_2024.xlsx"),
            os.path.join(home, "Downloads", "2024-IIAG-scores.xlsx"),
            os.path.join(home, "Downloads", "iiag_2024.xlsx"),
            os.path.join(home, "Desktop", "usaii", "2024-IIAG-scores.xlsx"),
            os.path.join(home, "Desktop", "usaii", "iiag_2024.xlsx"),
        ]
        for candidate in candidates:
            if os.path.exists(candidate):
                path = candidate
                print(f"  ✓ Found IIAG file: {path}")
                break
    if path is None:
        path = "2024-IIAG-scores.xlsx"  # fallback
    import os
    if not os.path.exists(path):
        print(f"\n  ⚠ IIAG file not found at '{path}'")
        print("  Download: https://iiag.online/downloads.html → '2024 IIAG: Data'")
        print("  Save as iiag_2024.xlsx next to this script. Skipping IIAG.\n")
        return {}
    try:
        import openpyxl
    except ImportError:
        print("  ⚠ Run: pip install openpyxl")
        return {}
    try:
        wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = None
        for name in ["Data", "IIAG Data", "Sheet1", wb.sheetnames[0]]:
            if name in wb.sheetnames:
                sheet = wb[name]; break
        rows    = list(sheet.iter_rows(values_only=True))
        headers = [str(h).strip() if h else "" for h in rows[0]]

        # IIAG 2024 Excel structure:
        # Row 0: IsVariable metadata
        # Row 1: Depth metadata  
        # Row 2: Level metadata
        # Row 3: SeriesID (column codes)
        # Row 4: Measure names
        # Row 5: Data source acronyms
        # Row 6: Header row (Country, Year, ...)
        # Row 7+: Data rows

        SERIES_ROW = 3   # SeriesID row index
        DATA_START = 7   # first data row
        COUNTRY_COL = 0
        YEAR_COL = 1

        # Pre-mapped column indices from 2024 IIAG Excel (stable across downloads)
        KNOWN_COL_IDX = {
            "GOVERNANCE": 2,   "SROL": 3,    "ROLJUS": 32,
            "ACCTRANS": 75,    "ANTICORR": 114, "PUBADMIN": 275,
            "BUSENV": 303,     "INFR": 331,  "HD": 369,
            "HEALTH": 370,     "EDUC": 406,  "FEO": 274,
        }

        # Verify against actual SeriesID row (self-healing if layout changes)
        series_row = rows[SERIES_ROW]
        col_idx = {}
        all_cols = {ind["col"] for inds in IIAG_INDICATORS.values() for ind in inds}
        for idx, val in enumerate(series_row):
            if val in all_cols:
                col_idx[val] = idx
        # Fall back to known indices for any not found
        for col, idx in KNOWN_COL_IDX.items():
            if col in all_cols and col not in col_idx:
                col_idx[col] = idx
                print(f"  ⚠ Using pre-mapped index for {col}")

        raw = {}
        target_countries = {"Nigeria", "Algeria"}
        for row in rows[DATA_START:]:
            if not row or row[COUNTRY_COL] not in target_countries: continue
            country = row[COUNTRY_COL]
            year = int(row[YEAR_COL]) if row[YEAR_COL] else 2023
            raw.setdefault(country, {}).setdefault(year, {})
            for col, idx in col_idx.items():
                if row[idx] is not None:
                    raw[country][year][col] = row[idx]

        # Map country name → iso3
        NAME_TO_ISO3 = {"Nigeria": "NGA", "Algeria": "DZA"}
        out = {}
        for country, years in raw.items():
            latest = max(years)
            iso3 = NAME_TO_ISO3.get(country, country)
            out[iso3] = {col: (float(val), str(latest)) for col, val in years[latest].items() if val is not None}
            print(f"  ✓ IIAG loaded {country} (year: {latest}) — {len(out[iso3])} indicators")
        return out
    except Exception as e:
        print(f"  ⚠ IIAG parse error: {e}")
        return {}


def merge_iiag(results, iiag_data):
    if not iiag_data: return results
    for iso2 in COUNTRIES:
        iso3 = IIAG_ISO3[iso2]
        country_iiag = iiag_data.get(iso3, {})
        for sector, indicators in IIAG_INDICATORS.items():
            results[iso2]["sectors"].setdefault(sector, {})
            for ind in indicators:
                col = ind["col"]
                if col in country_iiag:
                    value, year = country_iiag[col]
                    rec = {"id": ind["id"], "label": ind["label"], "value": float(value),
                           "year": year, "source": "iiag", "source_code": col,
                           "missing": False, "flag": _flag(year)}
                    stale = " ⚠ STALE" if rec["flag"] == "stale" else ""
                    print(f"  ✓ [IIAG] {rec['label']}: {rec['value']} ({year}){stale}")
                else:
                    rec = {"id": ind["id"], "label": ind["label"], "value": None,
                           "year": None, "source": None, "source_code": col,
                           "missing": True, "flag": "no_data"}
                    print(f"  ✗ MISSING [IIAG] {ind['label']}")
                results[iso2]["sectors"][sector][ind["id"]] = rec
    return results

if __name__ == "__main__":
    print(f"Multi-Source Fetch | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"Countries : {', '.join(m['name'] for m in COUNTRIES.values())}")
    print(f"Sources   : World Bank → ITU → UNESCO UIS → WHO GHO → Mo Ibrahim IIAG")
    print(f"Sectors   : {', '.join(INDICATORS.keys())}\n")

    results = fetch_all()

    print("\n── Mo Ibrahim IIAG ──────────────────────────────────────")
    iiag_data = load_iiag()  # auto-detects file location
    results   = merge_iiag(results, iiag_data)

    export_json(results)
    export_excel(results)
    print_summary(results)
