"""
Phase 2 — Scoring Engine | AI Readiness Copilot
Reads all_data.xlsx → normalizes indicators → computes sector scores with confidence levels
Output: scores.xlsx
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Normalization benchmarks ───────────────────────────────────────────────────
# Each indicator: (min, max, higher_is_better)
# min/max = realistic global range for developing economies
# higher_is_better = False means we invert (e.g. mortality: lower = better)

BENCHMARKS = {
    # EDUCATION
    "adult_literacy":       (30,  100, True),
    "edu_spend_gdp":        (1,   10,  True),
    "tertiary_enrollment":  (2,   80,  True),
    "primary_completion":   (40,  100, True),
    "trained_teachers":     (30,  100, True),
    "iiag_education":       (0,   100, True),
    "iiag_human_dev":       (0,   100, True),

    # WORKFORCE
    "labor_force":          (1e6, 200e6, True),   # absolute — lower weight
    "unemployment":         (2,   30,   False),   # lower = better
    "rd_spend":             (0,   3,    True),
    "ict_skills":           (5,   80,   True),
    "iiag_econ_opp":        (0,   100,  True),
    "iiag_business_env":    (0,   100,  True),

    # HEALTHCARE
    "health_spend_gdp":     (1,   12,   True),
    "hospital_beds":        (0.1, 5,    True),
    "under5_mortality":     (5,   150,  False),   # lower = better
    "physicians":           (0.1, 4,    True),
    "uhc_index":            (20,  90,   True),
    "iiag_health":          (0,   100,  True),

    # GOVERNMENT
    "iiag_overall":         (0,   100,  True),
    "iiag_srol":            (0,   100,  True),
    "iiag_rule_of_law":     (0,   100,  True),
    "iiag_accountability":  (0,   100,  True),

    # INFRASTRUCTURE
    "internet_users":       (5,   95,   True),
    "mobile_subs":          (20,  130,  True),
    "electricity_access":   (10,  100,  True),
    "broadband_subs":       (0,   40,   True),
    "mobile_broadband":     (5,   100,  True),
    "iiag_infra":           (0,   100,  True),
}

# ── Sector weights ─────────────────────────────────────────────────────────────
# Which indicators count toward each sector score, and how much
# Weights within a sector sum to 1.0

SECTOR_WEIGHTS = {
    "education": {
        "adult_literacy":      0.25,
        "tertiary_enrollment": 0.20,
        "primary_completion":  0.15,
        "edu_spend_gdp":       0.15,
        "trained_teachers":    0.15,
        "iiag_education":      0.10,
    },
    "workforce": {
        "unemployment":        0.25,
        "rd_spend":            0.20,
        "ict_skills":          0.20,
        "iiag_econ_opp":       0.20,
        "iiag_business_env":   0.15,
    },
    "healthcare": {
        "uhc_index":           0.30,
        "under5_mortality":    0.20,
        "physicians":          0.20,
        "health_spend_gdp":    0.15,
        "hospital_beds":       0.15,
    },
    "government": {
        "iiag_overall":        0.30,
        "iiag_srol":           0.25,
        "iiag_rule_of_law":    0.25,
        "iiag_accountability": 0.20,
    },
    "infrastructure": {
        "internet_users":      0.30,
        "electricity_access":  0.25,
        "broadband_subs":      0.20,
        "mobile_subs":         0.15,
        "iiag_infra":          0.10,
    },
}

# ── Confidence penalties ───────────────────────────────────────────────────────
PENALTY_MISSING = 0.20   # each missing indicator lowers confidence by 20%
PENALTY_STALE   = 0.08   # each stale indicator lowers confidence by 8%


# ── Normalization ──────────────────────────────────────────────────────────────

def normalize(ind_id, value):
    """Normalize a raw value to 0–100 using benchmarks."""
    if ind_id not in BENCHMARKS:
        return None
    lo, hi, higher_is_better = BENCHMARKS[ind_id]
    clamped = max(lo, min(hi, value))
    score = (clamped - lo) / (hi - lo) * 100
    return round(score if higher_is_better else 100 - score, 2)


# ── Load all_data.xlsx ─────────────────────────────────────────────────────────

def load_data(path="all_data.xlsx"):
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb["All Data"]
    rows = list(ws.iter_rows(values_only=True))

    headers = [str(h).strip() if h else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(headers)}

    data = {}   # {country_code: {sector: {ind_id: rec}}}

    for row in rows[1:]:
        if not row or row[col["Country Code"]] is None:
            continue
        iso2    = row[col["Country Code"]]
        sector  = row[col["Sector"]]
        ind_id  = row[col["Indicator ID"]]
        value   = row[col["Value"]]
        year    = row[col["Year"]]
        missing = str(row[col["Missing"]]).lower() == "true"
        flag    = row[col["Flag"]] or "ok"
        source  = row[col["Source"]] or ""

        data.setdefault(iso2, {}).setdefault(sector, {})[ind_id] = {
            "value": value, "year": year,
            "missing": missing, "flag": flag, "source": source,
        }

    return data


# ── Score one sector ───────────────────────────────────────────────────────────

def score_sector(sector, indicators, sector_data):
    """
    Returns (score, confidence, breakdown) for one sector.
    breakdown = list of {ind_id, label, raw, normalized, weight, weighted, flag, source}
    """
    weights    = SECTOR_WEIGHTS.get(sector, {})
    total_w    = sum(weights.values())
    breakdown  = []
    weighted_sum = 0.0
    used_weight  = 0.0
    missing_count = 0
    stale_count   = 0

    for ind_id, weight in weights.items():
        rec = sector_data.get(ind_id)

        if rec is None or rec["missing"] or rec["value"] is None:
            missing_count += 1
            breakdown.append({
                "ind_id": ind_id, "raw": None, "normalized": None,
                "weight": weight, "weighted": None,
                "flag": "missing", "source": "",
            })
            continue

        normalized = normalize(ind_id, float(rec["value"]))
        if normalized is None:
            missing_count += 1
            continue

        if rec["flag"] == "stale":
            stale_count += 1

        weighted_val  = normalized * weight
        weighted_sum += weighted_val
        used_weight  += weight

        breakdown.append({
            "ind_id":     ind_id,
            "raw":        rec["value"],
            "normalized": normalized,
            "weight":     weight,
            "weighted":   round(weighted_val, 3),
            "flag":       rec["flag"],
            "source":     rec["source"],
        })

    # Rescale to used weight (skip missing, don't penalize score directly)
    if used_weight > 0:
        score = round((weighted_sum / used_weight) * 100 / 100, 1)
    else:
        score = None

    # Confidence: start at 100, penalize missing + stale
    confidence = 100
    confidence -= missing_count * PENALTY_MISSING * 100
    confidence -= stale_count   * PENALTY_STALE   * 100
    confidence  = max(0, round(confidence))

    return score, confidence, breakdown


# ── Score all countries ────────────────────────────────────────────────────────

def score_all(data):
    results = {}
    for iso2, sectors in data.items():
        results[iso2] = {"sectors": {}, "overall": None}
        sector_scores = []

        for sector in SECTOR_WEIGHTS:
            sector_data = sectors.get(sector, {})
            score, confidence, breakdown = score_sector(sector, SECTOR_WEIGHTS[sector], sector_data)
            results[iso2]["sectors"][sector] = {
                "score": score,
                "confidence": confidence,
                "breakdown": breakdown,
            }
            if score is not None:
                sector_scores.append(score)

        # Overall = average of sector scores
        if sector_scores:
            results[iso2]["overall"] = round(sum(sector_scores) / len(sector_scores), 1)

    return results


# ── Export scores.xlsx ─────────────────────────────────────────────────────────

def export_scores(results, country_names, path="scores.xlsx"):
    wb = openpyxl.Workbook()

    # Styles
    hdr_fill  = PatternFill("solid", fgColor="1F3864")
    ng_fill   = PatternFill("solid", fgColor="008751")   # Nigeria green
    dz_fill   = PatternFill("solid", fgColor="006233")   # Algeria green
    hi_fill   = PatternFill("solid", fgColor="E2EFDA")
    mid_fill  = PatternFill("solid", fgColor="FFF2CC")
    lo_fill   = PatternFill("solid", fgColor="FCE4D6")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    bold      = Font(bold=True)
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def sc(cell, fill=None, font=None, align="center"):
        if fill: cell.fill = fill
        if font: cell.font = font
        else: cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = border

    def score_fill(s):
        if s is None: return lo_fill
        return hi_fill if s >= 60 else (mid_fill if s >= 40 else lo_fill)

    def conf_label(c):
        if c >= 80: return "HIGH"
        if c >= 60: return "MEDIUM"
        return "LOW ⚠"

    # ── Sheet 1: Sector Scorecard ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Sector Scorecard"

    headers = ["Sector", "Nigeria Score", "Nigeria Confidence",
               "Algeria Score", "Algeria Confidence", "Gap (DZ - NG)"]
    ws.append(headers)
    for i in range(1, 7):
        sc(ws.cell(1, i), fill=hdr_fill, font=hdr_font)

    row = 2
    for sector in SECTOR_WEIGHTS:
        ng = results.get("NG", {}).get("sectors", {}).get(sector, {})
        dz = results.get("DZ", {}).get("sectors", {}).get(sector, {})
        ng_s = ng.get("score")
        dz_s = dz.get("score")
        gap  = round(dz_s - ng_s, 1) if ng_s is not None and dz_s is not None else None

        ws.append([
            sector.upper(),
            ng_s, f"{ng.get('confidence', 0)}% ({conf_label(ng.get('confidence', 0))})",
            dz_s, f"{dz.get('confidence', 0)}% ({conf_label(dz.get('confidence', 0))})",
            gap,
        ])
        sc(ws.cell(row, 1), font=bold, align="left")
        sc(ws.cell(row, 2), fill=score_fill(ng_s))
        sc(ws.cell(row, 3))
        sc(ws.cell(row, 4), fill=score_fill(dz_s))
        sc(ws.cell(row, 5))
        gap_fill = hi_fill if gap and gap > 0 else lo_fill
        sc(ws.cell(row, 6), fill=gap_fill if gap else None)
        row += 1

    # Overall row
    ng_ov = results.get("NG", {}).get("overall")
    dz_ov = results.get("DZ", {}).get("overall")
    ws.append(["OVERALL AI READINESS", ng_ov, "", dz_ov, "",
               round(dz_ov - ng_ov, 1) if ng_ov and dz_ov else None])
    for i in range(1, 7):
        sc(ws.cell(row, i), fill=hdr_fill, font=hdr_font)
    ws.cell(row, 2).fill = score_fill(ng_ov)
    ws.cell(row, 4).fill = score_fill(dz_ov)

    widths = [18, 16, 22, 16, 22, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ── Sheet 2: Indicator Breakdown ───────────────────────────────────────────
    ws2 = wb.create_sheet("Indicator Breakdown")
    hdrs2 = ["Sector", "Indicator", "Weight",
             "NG Raw", "NG Normalized", "NG Flag",
             "DZ Raw", "DZ Normalized", "DZ Flag"]
    ws2.append(hdrs2)
    for i in range(1, 10):
        sc(ws2.cell(1, i), fill=hdr_fill, font=hdr_font)

    row2 = 2
    sect_fill_map = PatternFill("solid", fgColor="2E75B6")
    for sector in SECTOR_WEIGHTS:
        # Sector header
        ws2.append([sector.upper()] + [""] * 8)
        for i in range(1, 10):
            sc(ws2.cell(row2, i), fill=sect_fill_map,
               font=Font(bold=True, color="FFFFFF", size=10))
        row2 += 1

        ng_bd = {b["ind_id"]: b for b in results.get("NG", {}).get("sectors", {}).get(sector, {}).get("breakdown", [])}
        dz_bd = {b["ind_id"]: b for b in results.get("DZ", {}).get("sectors", {}).get(sector, {}).get("breakdown", [])}

        for ind_id, weight in SECTOR_WEIGHTS[sector].items():
            ng_b = ng_bd.get(ind_id, {})
            dz_b = dz_bd.get(ind_id, {})
            ws2.append([
                "", ind_id, f"{int(weight*100)}%",
                ng_b.get("raw"), ng_b.get("normalized"), ng_b.get("flag", ""),
                dz_b.get("raw"), dz_b.get("normalized"), dz_b.get("flag", ""),
            ])
            for i in range(1, 10):
                f = None
                if i in (5,):  f = score_fill(ng_b.get("normalized"))
                if i in (8,):  f = score_fill(dz_b.get("normalized"))
                if i == 6 and ng_b.get("flag") == "stale": f = mid_fill
                if i == 9 and dz_b.get("flag") == "stale": f = mid_fill
                sc(ws2.cell(row2, i), fill=f, align="left" if i <= 2 else "center")
            row2 += 1

    for col, w in zip(["A","B","C","D","E","F","G","H","I"],
                       [14, 22, 8, 16, 14, 10, 16, 14, 10]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Confidence flags ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Confidence Flags")
    ws3.append(["Country", "Sector", "Score", "Confidence %",
                "Confidence Level", "Action Required"])
    for i in range(1, 7):
        sc(ws3.cell(1, i), fill=hdr_fill, font=hdr_font)

    row3 = 2
    country_map = {"NG": "Nigeria", "DZ": "Algeria"}
    for iso2 in ["NG", "DZ"]:
        for sector in SECTOR_WEIGHTS:
            s = results.get(iso2, {}).get("sectors", {}).get(sector, {})
            score = s.get("score")
            conf  = s.get("confidence", 0)
            level = conf_label(conf)
            action = "✓ OK" if conf >= 80 else (
                "⚠ Supplement with NBS/ONS data" if conf >= 60 else
                "🔴 Human expert review required"
            )
            ws3.append([country_map[iso2], sector.upper(), score, conf, level, action])
            row_fill = hi_fill if conf >= 80 else (mid_fill if conf >= 60 else lo_fill)
            for i in range(1, 7):
                sc(ws3.cell(row3, i), fill=row_fill, align="left" if i in (1,2,5,6) else "center")
            row3 += 1

    for col, w in zip(["A","B","C","D","E","F"], [12,16,10,14,14,32]):
        ws3.column_dimensions[col].width = w

    wb.save(path)
    print(f"✓ Scores → {path}  (sheets: Sector Scorecard | Indicator Breakdown | Confidence Flags)")


# ── Print summary ──────────────────────────────────────────────────────────────

def print_scores(results, country_names):
    print("\n" + "="*55)
    print("AI READINESS SCORES")
    print("="*55)
    for iso2, name in country_names.items():
        print(f"\n{name}")
        print(f"  Overall AI Readiness: {results[iso2]['overall']}/100")
        for sector, data in results[iso2]["sectors"].items():
            score = data["score"]
            conf  = data["confidence"]
            bar   = "█" * int((score or 0) // 5)
            print(f"  {sector:<16} {str(score):<6} {bar:<20} confidence: {conf}%")


# ── Run ────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Phase 2 — Scoring Engine")
    print("Reading all_data.xlsx ...\n")

    data    = load_data("all_data.xlsx")
    results = score_all(data)

    country_names = {
        iso2: info.get("name", iso2)
        for iso2, info in data.items()
    } if data else {"NG": "Nigeria", "DZ": "Algeria"}

    # Fallback names
    country_names.setdefault("NG", "Nigeria")
    country_names.setdefault("DZ", "Algeria")

    print_scores(results, country_names)
    export_scores(results, country_names, "scores.xlsx")
